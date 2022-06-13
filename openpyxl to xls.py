from openpyxl import  *
from tkinter import *
 wb = load_workbook('excel.xlsx')
 sheet = wb.active
 def excel():
     sheet.column_dimensions['A'].width = 30
     sheet.column_dimensions['B'].width = 10
     sheet.column_dimensions['C'].width = 10
     sheet.column_dimensions['D'].width = 20
     sheet.column_dimensions['E'].width = 30
     sheet.column_dimensions['F'].width = 40
     sheet.column_dimensions['G'].width = 50

     # write given data to an excel spreadsheet
     #at particular location
     sheet.cell(row=1, column=1).value = "Name"
sheet.cell(row=1, column=2).value = "Course"
sheet.cell(row=1, column=3).value = "Semester"
sheet.cell(row=1, column=4).value = "Form Number"
sheet.cell(row=1, column=5).value = "Contact Number"
sheet.cell(row=1, column=6).value = "Email ID"
sheet.cell(row=1, column=7).value = "Address"


# function to set foruc
def focus1 (event):
    course_field.focus_set()
def focus2(event):
    sem_field.focus_set()
def focus3 (event):
    form_field.focus_set()
def focus4 (event):
    Con_field.focus_set()
def focus5 (event):
    Email_id_field.focus_set()
def focus6(event):
    Address_field.focus_set()

def clear():
    name_field.delete(0, END)
    course_field.delete(0,END)
    sem_field.delete(0,END)
    form_field.delete(0,END)
    Con_field.delete(0,END)
    Email_Id_field.delete(0,END)
    Address_field.delete(0,END)

def insert():
    if (name_field.get() == "" and
    course_field.get() == "" and
    sem_field.get() == "" and
    form_field.get() == "" and
    Con_field.get () == "" and
    Email_Id_field.get () == "" and
    Address_field.get () == ""):

print("empty input")
else:

sheet.cell(row=current_raw + 1,column=1).value = name_field.get()
sheet.cell(row=current_raw + 1,column=2).value = course_field.get()
sheet.cell(row=current_raw + 1,column=3).value = sem_field.get()
sheet.cell(row=current_raw + 1,column=4).value = form_field.get()
sheet.cell(row=current_raw + 1,column=5).value = Con_field.get()
sheet.cell(row=current_raw + 1,column=6).value = Email_Id_field.get()
sheet.cell(row=current_raw + 1,column=7).value = address_field.get()

    wb.save('excel.xlsx')
    name_field.focus.set()
    clear()
if __name__ == "__main__"
    root = Tk()
    root.configure(background='blue')

    root.title("Regristration Form")
    root.geometry("500x300")
    excel()

    heading = Label(root, text="Form", bg="Light green")
    name = Label(root,text="Name",bg="Light green")
    course = Label(root, text="Course",bg="Light green")
    sem = Label(root, text="semester", bg="Light green")
    form = Label(root, text="Form Number", bg="Light green")
    Con = Label(root, text="Contact Number", bg="Light green")
    Email_Id = Label(root,text="Email ID", bg="light green")
    address = Label(root, text="address", bg="light green")



    heading.grid(row=0, column=1)

    name.grid(row=1, column=0)
    course.grid(row=2,column=0)
    sem.grid(row=3,column=0)
    form.grid(row=4,column=0)
    Con.grid(row=5,column=0)
    Email_Id.grid(row=6,column=0)
    address.grid(row=7,column=0)

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_field = Entry(root)
    Con_field = Entry(root)
    Email_Id_field = Entry(root)
    address_field = Entry(root)

    name_field.bind()






